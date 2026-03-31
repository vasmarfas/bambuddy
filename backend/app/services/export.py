import csv
import io
from datetime import datetime
from typing import Any

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from backend.app.models.archive import PrintArchive


class ExportService:
    """Service for exporting archive data to CSV/Excel formats."""

    # Default fields to export
    DEFAULT_FIELDS = [
        "id",
        "print_name",
        "filename",
        "status",
        "quantity",
        "printer_id",
        "project_name",
        "filament_type",
        "filament_used_grams",
        "print_time_seconds",
        "layer_height",
        "nozzle_diameter",
        "bed_temperature",
        "nozzle_temperature",
        "total_layers",
        "cost",
        "designer",
        "tags",
        "notes",
        "failure_reason",
        "started_at",
        "completed_at",
        "created_at",
    ]

    # Field labels for headers
    FIELD_LABELS = {
        "id": "ID",
        "print_name": "Print Name",
        "filename": "Filename",
        "status": "Status",
        "quantity": "Items Printed",
        "printer_id": "Printer ID",
        "project_name": "Project",
        "filament_type": "Filament Type",
        "filament_used_grams": "Filament (g)",
        "print_time_seconds": "Print Time (s)",
        "layer_height": "Layer Height (mm)",
        "nozzle_diameter": "Nozzle (mm)",
        "bed_temperature": "Bed Temp (°C)",
        "nozzle_temperature": "Nozzle Temp (°C)",
        "total_layers": "Total Layers",
        "cost": "Cost",
        "designer": "Designer",
        "tags": "Tags",
        "notes": "Notes",
        "failure_reason": "Failure Reason",
        "started_at": "Started At",
        "completed_at": "Completed At",
        "created_at": "Created At",
    }

    def __init__(self, db: AsyncSession):
        self.db = db

    async def export_archives(
        self,
        format: str = "csv",
        fields: list[str] | None = None,
        printer_id: int | None = None,
        project_id: int | None = None,
        status: str | None = None,
        date_from: datetime | None = None,
        date_to: datetime | None = None,
        search: str | None = None,
    ) -> tuple[bytes, str, str]:
        """Export archives to CSV or Excel format.

        Args:
            format: Export format ('csv' or 'xlsx')
            fields: List of fields to include (None = all default fields)
            printer_id: Filter by printer
            project_id: Filter by project
            status: Filter by status
            date_from: Filter by start date
            date_to: Filter by end date
            search: Search filter

        Returns:
            Tuple of (file_bytes, filename, content_type)
        """
        # Build query
        query = (
            select(PrintArchive).options(selectinload(PrintArchive.project)).order_by(PrintArchive.created_at.desc())
        )

        # Apply filters
        if printer_id:
            query = query.where(PrintArchive.printer_id == printer_id)
        if project_id:
            query = query.where(PrintArchive.project_id == project_id)
        if status:
            query = query.where(PrintArchive.status == status)
        if date_from:
            query = query.where(PrintArchive.created_at >= date_from)
        if date_to:
            query = query.where(PrintArchive.created_at <= date_to)
        if search:
            like_pattern = f"%{search}%"
            query = query.where(
                (PrintArchive.print_name.ilike(like_pattern))
                | (PrintArchive.filename.ilike(like_pattern))
                | (PrintArchive.tags.ilike(like_pattern))
                | (PrintArchive.notes.ilike(like_pattern))
                | (PrintArchive.designer.ilike(like_pattern))
            )

        # Execute query
        result = await self.db.execute(query)
        archives = list(result.scalars().all())

        # Determine fields to export
        export_fields = fields if fields else self.DEFAULT_FIELDS

        # Convert to rows
        rows = []
        for archive in archives:
            row = self._archive_to_row(archive, export_fields)
            rows.append(row)

        # Generate headers
        headers = [self.FIELD_LABELS.get(f, f) for f in export_fields]

        # Generate file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        if format == "xlsx":
            file_bytes = self._generate_xlsx(headers, rows, export_fields)
            filename = f"archives_export_{timestamp}.xlsx"
            content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        else:
            file_bytes = self._generate_csv(headers, rows)
            filename = f"archives_export_{timestamp}.csv"
            content_type = "text/csv"

        return file_bytes, filename, content_type

    async def export_stats(
        self,
        format: str = "csv",
        days: int = 30,
        printer_id: int | None = None,
        project_id: int | None = None,
        created_by_id: int | None = None,
    ) -> tuple[bytes, str, str]:
        """Export statistics summary to CSV or Excel format.

        Args:
            format: Export format ('csv' or 'xlsx')
            days: Number of days to include in stats
            printer_id: Filter by printer
            project_id: Filter by project
            created_by_id: Filter by user who created the print (-1 for no user)

        Returns:
            Tuple of (file_bytes, filename, content_type)
        """
        from backend.app.services.failure_analysis import FailureAnalysisService

        # Get failure analysis data (includes stats)
        analysis_service = FailureAnalysisService(self.db)
        analysis = await analysis_service.analyze_failures(
            days=days,
            printer_id=printer_id,
            project_id=project_id,
            created_by_id=created_by_id,
        )

        # Build stats rows
        rows = [
            ["Metric", "Value"],
            ["Period (days)", analysis["period_days"]],
            ["Total Prints", analysis["total_prints"]],
            ["Failed Prints", analysis["failed_prints"]],
            ["Failure Rate (%)", analysis["failure_rate"]],
            [""],
            ["Failures by Reason", ""],
        ]

        for reason, count in analysis["failures_by_reason"].items():
            rows.append([reason, count])

        rows.append([""])
        rows.append(["Failures by Filament", ""])

        for filament, count in analysis["failures_by_filament"].items():
            rows.append([filament, count])

        rows.append([""])
        rows.append(["Failures by Printer", ""])

        for printer, count in analysis["failures_by_printer"].items():
            rows.append([printer, count])

        rows.append([""])
        rows.append(["Weekly Trend", ""])
        rows.append(["Week", "Total", "Failed", "Rate (%)"])

        for week in analysis["trend"]:
            rows.append(
                [
                    week["week_start"],
                    week["total_prints"],
                    week["failed_prints"],
                    week["failure_rate"],
                ]
            )

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        if format == "xlsx":
            file_bytes = self._generate_xlsx_simple(rows)
            filename = f"stats_export_{timestamp}.xlsx"
            content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        else:
            file_bytes = self._generate_csv_simple(rows)
            filename = f"stats_export_{timestamp}.csv"
            content_type = "text/csv"

        return file_bytes, filename, content_type

    def _archive_to_row(self, archive: PrintArchive, fields: list[str]) -> list[Any]:
        """Convert an archive to a row of values."""
        row = []
        for field in fields:
            if field == "project_name":
                value = archive.project.name if archive.project else None
            elif field in ("started_at", "completed_at", "created_at"):
                value = getattr(archive, field)
                if value:
                    value = value.isoformat()
            else:
                value = getattr(archive, field, None)
            row.append(value)
        return row

    def _generate_csv(self, headers: list[str], rows: list[list]) -> bytes:
        """Generate CSV file content."""
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        writer.writerows(rows)
        return output.getvalue().encode("utf-8")

    def _generate_csv_simple(self, rows: list[list]) -> bytes:
        """Generate CSV file content from simple rows (no separate headers)."""
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerows(rows)
        return output.getvalue().encode("utf-8")

    def _generate_xlsx(self, headers: list[str], rows: list[list], fields: list[str]) -> bytes:
        """Generate Excel file content."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

        wb = Workbook()
        ws = wb.active
        ws.title = "Archives"

        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center")

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Write data
        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-adjust column widths
        for col_idx, _field in enumerate(fields, 1):
            column_letter = get_column_letter(col_idx)
            max_length = len(headers[col_idx - 1])
            for row in rows:
                cell_value = row[col_idx - 1]
                if cell_value is not None:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        # Freeze header row
        ws.freeze_panes = "A2"

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    def _generate_xlsx_simple(self, rows: list[list]) -> bytes:
        """Generate Excel file content from simple rows."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
        except ImportError:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

        wb = Workbook()
        ws = wb.active
        ws.title = "Statistics"

        bold_font = Font(bold=True)

        for row_idx, row in enumerate(rows, 1):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                # Bold section headers
                if col_idx == 1 and value and isinstance(value, str) and value.endswith(":"):
                    cell.font = bold_font

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
